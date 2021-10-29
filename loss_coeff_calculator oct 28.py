#loss_coeff_calculator.py
import numpy as np
import pandas as pd
import os
from scipy.signal import find_peaks
from math import log
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import get_column_letter
from scipy.fft import fft,fftfreq
from scipy import fftpack
import glob
from time import sleep

Folder_Directory=r'C:\Users\proberts\Desktop\GAH2107-09'
#file_name_tosave=r'U:\7114\Copy of NOV20-GAH2025-05-3.3-FP2_1600_2.xlsx'
#chiplength=4.63 #cm
n_eff=2.1361
os.chdir(Folder_Directory)
for file in glob.glob("*.dat"):
     #What file is being converted
     print(file) 
     df = pd.read_table(file,delimiter=',',header=None,index_col=False)
     #print(df)
     replacedf=df.iloc[8:].astype(float)
     df.iloc[8:]=replacedf
     file1 = file.replace('dat','xlsx')
     if os.path.exists(file1):
         os.remove(file1)
         print('removed and replaced file:'+file1)
     else:
         print('making file')
     df.to_excel(file1,index=False,header=False)
print('done converting dat')

###Set up where to look for the data to manipulate
files=os.listdir(Folder_Directory)
print(files)
for item in files:
    if item.endswith('.xlsx'):
        file_name_tosave=Folder_Directory+'\\' +item
        print(file_name_tosave)
        
        

        
        data=pd.read_excel(Folder_Directory+'\\'+item,header=None,engine='openpyxl',index_col=False)
        #data=[data[8:][0]]+[data[8:][2]]'

        wavelength=data[8:][0].to_numpy()
        correctedoutpower=data[8:][2].to_numpy()

        chiplengthraw=data[3][0]
        chiplength=chiplengthraw.replace('Length (cm) = ','')
        chiplength=chiplength.replace('cmhorizontal polarization','')
        chiplength=float(chiplength)
        ###Read The columns of the file
        #data.columns=[
        #    "# Wavelength (nm)",
        #    " Corrected Output Power (mW)"
        #]
        
        ###Perfom FFT and DCT signal processing

        ndata=[wavelength]+[correctedoutpower]
        tndata=np.array(ndata)
        N=len(tndata[1])
        T=1
        fftsig=fft(tndata[1])
        dctsig=fftpack.dct(tndata[1],norm='ortho')
        dctsig[60:]=0
        smoothed=fftpack.idct(dctsig,norm='ortho')
        freqfft=fftfreq(N,T)[:N//2]
        dominating_distance_index=np.argmax(2.0/N * np.abs(fftsig[1:N//2]))+1
        dominating_distance=1/freqfft[dominating_distance_index]#finds the dominating period with which the data is goverend by
        #print('the period of the data is',dominating_distance)

        ####Plot FFT data
        fig,axs=plt.subplots(2,1)

        #axs[0] gives you the frequency
        axs[0].plot(freqfft,2.0/N * np.abs(fftsig[0:N//2]))
        axs[0].plot(freqfft[np.argmax(np.abs(fftsig[1:N//2]))+1],np.max(2.0/N * np.abs(fftsig[1:N//2])),'*')
        axs[0].set_xlim(0,np.max(freqfft)/2)
        axs[0].set_xlabel('Frequency (1/steps)')
        axs[0].set_ylabel('Amplitude')
        axs[0].grid(True)

        #axs[1] gives you the period which is more useful for this code
        axs[1].plot(1/freqfft[1:],2.0/N * np.abs(fftsig[1:N//2]))
        axs[1].set_xlim(0,200)
        axs[1].set_xlabel('Period of the data (steps)')
        axs[1].set_ylabel('Amplitude')
        axs[1].grid(True)

        fig.suptitle('FFT of Data',fontsize='16')
        fig.tight_layout()

        fftfigname='FFT of Data.png'
        if os.path.exists(fftfigname):
            os.remove(fftfigname)
        plt.savefig(fftfigname)
        #plt.show()
        plt.close()

        ###Finding locations of the peaks and the roughs i the data 
        loc_peaks, _ = find_peaks(tndata[1], distance=int(.65*dominating_distance)) #Check the .9 value if the peaks arent being captured correctly
        loc_troughs, _=find_peaks(1/tndata[1],distance=int(.65*dominating_distance))
        
        
        #logic to organize the peaks and troughs in the order in which the occur
        wlpeaks=[]
        powpeaks=[]
        wltroughs=[]
        powtroughs=[]
        for i in loc_peaks:
            wlpeaks.append(tndata[0][i])
            powpeaks.append(tndata[1][i])
        for i in loc_troughs:
            wltroughs.append(tndata[0][i])
            powtroughs.append(tndata[1][i])       
        peaks=[wlpeaks]+[powpeaks]
        peaks=np.array(peaks).transpose().tolist()
        troughs=[wltroughs]+[powtroughs]
        troughs=np.array(troughs).transpose().tolist()
        list1=peaks+troughs 
        
        def sortfirst(val):
            return val[0]

        list1.sort(key=sortfirst)#list1 gives the peaks troughs and there locations
        list1=np.array(list1).transpose().tolist()

        locations=np.concatenate((loc_troughs,loc_peaks))
        
        locations=locations.tolist()
        print(locations)
        locations.sort()
        print(locations)

        if locations[1]-locations[0]<0.3*dominating_distance:
            del list1[0][0]
            del list1[1][0]

        if locations[len(locations)-1]-locations[len(locations)-2]<0.4*dominating_distance:
            del list1[0][len(list1[0])-1]
            del list1[1][len(list1[1])-1]
        ###Everyother average averages the P2P values and T2T values of the next peak or trough neighbor
        every_other_average=[]
        for i in range(len(list1[0])-2):
            every_other_average.append((list1[1][i]+list1[1][i+2])/2)
        

        #Contrast caluclation spelt with a K
        kontrast=[]
        #first value
        kontrast.append(abs(list1[1][0]-list1[1][1])/(list1[1][0]+list1[1][1]))
        #middle values
        for i in range(len(list1[1])-2):
            kontrast.append(abs(list1[1][i+1]-every_other_average[i])/(list1[1][i+1]+every_other_average[i]))
        #last value
        kontrast.append(abs(list1[1][len(list1[1])-2]-list1[1][len(list1[1])-1])/(list1[1][len(list1[1])-2]+list1[1][len(list1[1])-1]))
        

        #loss reflection ~R
        loss_Reflection=[]
        for K in kontrast:
            R_tilda=(1/K)*(1-(1-K**2)**.5)
            loss_Reflection.append(R_tilda)
        

        #attenuation coefficient alpha
        R=((n_eff-1)/(n_eff+1))**2
        alpha=[]
        for R_tilda in loss_Reflection:
            alpha.append(4.34/chiplength*(log(R)-log(R_tilda)))
        
        ####Ploting the Data,Smoothed data by doing dct and removing high frequncy values and than idct, and Peaks/troughs
        plt.plot(tndata[0],tndata[1],'.',label='Raw Data')
        plt.plot(tndata[0],smoothed,'--',label='Smoothed')
        plt.plot(list1[0],list1[1],'*',label='Local Extrema',mew=10)
        plt.ylim((0,1.2*np.max(every_other_average)))
        plt.legend(loc=4)
        plt.title('FP Measurement')
        plt.xlabel('Wavelength (nm)')
        plt.ylabel('Transmitted Power (mW)')
        plt.grid()

        fpfigname='FP Measurement.png'
        if os.path.exists(fpfigname):
            os.remove(fpfigname)
        plt.savefig(fpfigname)
        #plt.show()
        plt.close()

        #### adding another plot for alpha
        plt.plot(list1[0],alpha,'*')
        plt.title('Loss Coefficient')
        plt.ylabel('Loss Coefficient (dBm/cm)')
        plt.xlabel('Wavelength (nm)')
        plt.ylim((0,np.max(alpha)*1.2))
        plt.grid()

        alphafig='Loss Coefficient.png'
        if os.path.exists(alphafig):
            os.remove(alphafig)
        plt.savefig(alphafig)
        #plt.show()
        plt.close()

        ####Average and std dev of alpha calc
        Average_alpha=np.mean(alpha)
        Stddev_alpha=np.std(alpha)

        #print('average',Average_alpha)
        #print('standard dev',Stddev_alpha)

        ####re aranging data and placing values into an excel file
        every_other_average[:0]=[0]
        every_other_average.append(0)
        combining_calculations=[list1[0]]+[list1[1]]+[every_other_average]+[kontrast]+[loss_Reflection]+[alpha]
        combining_calculations=np.array(combining_calculations).transpose()
        #print(combining_calculations)

        df=pd.DataFrame(combining_calculations)


        df.columns=['Wavelength (nm)','Peak/Trough Power (mW)','P2P/T2T average (mW)','Contrast [K]','Loss Reflection Factor [~R]','Attenuation Coefficient [alpha] (dB/cm)']
        df['Wavelength (nm)']=df['Wavelength (nm)'].astype(float)
        #print(df)
        
        file_name=file_name_tosave
        #writer = pd.ExcelWriter(file_name, engine='openpyxl')

        #existing_data=pd.read_excel(file_name)

        #print(df)
        wb=openpyxl.load_workbook(file_name)


        with pd.ExcelWriter(file_name,engine='openpyxl') as writer:
            writer.book = wb
            
            writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
            df.to_excel(writer, 'loss calculations', index=False)
            ws=wb['loss calculations']
            column_towrite=len(combining_calculations[0])+1
            ws.cell(row=2,column=column_towrite,value="Average Loss Coeff")
            ws.cell(row=3,column=column_towrite,value='Standard Dev')
            ws.cell(row=2,column=column_towrite+1,value=Average_alpha)
            ws.cell(row=3,column=column_towrite+1,value=Stddev_alpha)

            img1=Image(fpfigname)
            img2=Image(fftfigname)
            img3=Image(alphafig)
            ws.add_image(img1,get_column_letter(column_towrite+2)+'2')
            ws.add_image(img2,get_column_letter(column_towrite+2)+'26')
            ws.add_image(img3,get_column_letter(column_towrite+12)+'2')
            
            wb.save(file_name)
            wb.close()
            
            
            